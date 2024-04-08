# -*- coding: utf-8 -*-
"""
Created on Wed Dec 21 15:02:56 2022

@author: camer
"""

import pandas as pd
from itertools import product
import openpyxl
import numpy as np
import matplotlib.pyplot as plt
from matplotlib import colors
import matplotlib.patches as mpatches
import math


def socialFitness(land_code, chromosome, transportation, n_cells):
    
    social_fit = {'comm_dist':[], 'transport_dist':[], 'rec_dist':[], 'fire_dist':[], 'health_dist':[]}
    
    residential = chromosome[5]
    commercial = chromosome[9]
    recreation = chromosome[6]
    fire = land_code[8]
    health = land_code[7]
    flood = []
    
    for r in residential:
        
        flood.append(n_cells[r[0]][r[1]])
        
        comm_d = np.mean([math.dist(r,c) for c in commercial])
        social_fit['comm_dist'].append(comm_d)
        
        rec_d = np.mean([math.dist(r,rec) for rec in recreation])
        social_fit['rec_dist'].append(rec_d)
        
        tr_d = np.mean([math.dist(r,t) for t in transportation])
        social_fit['transport_dist'].append(tr_d)
        
        fire_d = np.mean([math.dist(r,f) for f in fire])
        social_fit['fire_dist'].append(fire_d)
        
        health_d = np.mean([math.dist(r,h) for h in health])
        social_fit['health_dist'].append(health_d)
        
    flood_ratio = 1 - sum(flood)/len(flood)
    # print(flood_ratio)
    social_fit = {k:np.mean(v) for k,v in social_fit.items()}
    
    score = np.mean([1-((el-10)/10) for el in social_fit.values()]+[flood_ratio])
    
    return social_fit, score

def commFitness(chromosome, transportation):
    comm_fit = {'transport_dist':[], "pedestrian_traffic":[]}
    commercial = chromosome[9]
    
    wash_st = [(i,25) for i in range(15)]
    
    for c in commercial:
    
        tr_d = np.mean([math.dist(c,t) for t in transportation])
        comm_fit['transport_dist'].append(tr_d)
        
        traffic = np.mean([math.dist(c,w) for w in wash_st])
        comm_fit['pedestrian_traffic'].append(traffic)
        
    rows, cols = list(zip(*commercial))
    ratio = [(len(set(rows))/len(rows)) , (len(set(cols))/len(cols))]
    
    comm_fit = {k:np.mean(v) for k,v in comm_fit.items()}

    comm_density = min(ratio) / max(ratio)
    # print(comm_density)
    # comm_fit['density'] = 1-comm_density
    comm_vals = list(comm_fit.values())
    # print(comm_vals)
    score = [1-((el-10)/10) for el in comm_vals]+[1-comm_density]
    score = np.mean([i if i > 0 else 0 for i in score])
    
    return comm_fit, score

def envFitness(chromosome, flood_ratio, center_pt, n_cells):
    env_fit = {'density':[], 'industrial_dist':[], 'ind_res':[]}
    environmental = chromosome[6]
    # waterfront = chromosome[10]
    industrial = chromosome[4]
    residential = chromosome[5]
    
    flood = 0
    to_center = []
    for e in environmental:
        ind_d = np.mean([math.dist(e,i) for i in industrial])
        to_center.append(math.dist(e,center_pt))
        a,b = e
        if n_cells[a][b] == 1: flood += 1
        
        env_fit['industrial_dist'].append(ind_d)
    
    if np.mean(to_center) < 10: 
        env_fit['density'] = 1
    else: 
        env_fit['density'] = 1-((to_center-10)/10)
            
    for r in residential:
        ind_res = np.mean([math.dist(r,i) for i in industrial])
        env_fit['ind_res'].append(ind_res)
    
    flood_mit = flood/len(environmental)
    env_fit['flood_mitigation'] = (flood_ratio-flood_mit)/flood_ratio
    # want to maximize dist to industrial areas
    env_fit['industrial_dist'] = (np.mean(env_fit['industrial_dist'])-10)/10
    env_fit['ind_res'] = (np.mean(env_fit['ind_res'])-10)/10
    
    score = np.mean(list(env_fit.values()))

    return env_fit, score


def plotData(data, l_cells):
    
    data = np.array(data)
    
    fig = plt.figure(0, figsize=(5,5))
    plt.title('Hoboken Distribution of Land by Type')
    plt.box(False)
    
    cmap = colors.ListedColormap(["green","blue",'red','orange','grey','cyan','magenta','brown', 'yellow', 'pink'])
    plt.imshow(data, cmap = cmap)
    
    ax = fig.gca()
    ax.set_xticks(np.arange(0, 14, 2))
    ax.set_yticks(np.arange(0, 28, 2))
    
    # plt.grid(True,color='white')
    plt.axis([0,14, 0,28])
    
    my_colors = {
        'Ignore': 0,            'Recreation': 6,
        'Water': 1,             'Hospital': 7,
        'Undecided': 2,         'Fire Station': 8,
        'University': 3,        'Commercial': 9,
        'Industrial': 4,        'Waterfront': 10,
        'Residential': 5
                 }
    
    patches = [mpatches.Patch(color=cmap(v), label=k) for k,v in my_colors.items()]
    plt.legend(handles=patches, loc=2, bbox_to_anchor=(1.01,1))
    
    plt.show()
    
    def plotMap(l_cells):
        plt.imshow(np.array(l_cells), cmap = 'plasma')
        plt.colorbar()
    
    plotMap(l_cells)
    
    
def initVars(fpath):
    wb = openpyxl.load_workbook(fpath)
    flood_layer = wb['n']
    land_layer = wb['l']
    crossover_inclusion = wb['i']
    
    transportation = [(1,5),(2,4),(2,5),(3,4),(3,5)] + [(i,0) for i in range(28,37)] + [(i,1) for i in range(28,36)]
    
    n_cells = [] # flood risk high/low
    for rows in flood_layer.iter_rows():
        row_cells = []
        for cell in rows:
            row_cells.append(cell.value)
        n_cells.append(tuple(row_cells))
    
    flood_ratio = [item for n in n_cells for item in n]
    flood_ratio = [n for n in flood_ratio if n in [0,1]]
    flood_ratio = len(np.nonzero(flood_ratio)[0])/len(flood_ratio) # high risk/ low risk
    
    idx = len(n_cells)
    cols = len(n_cells[0])
    
    center_pt = (idx//2, cols//2)
    
    l_cells = [] # land code 0-10
    for rows in land_layer.iter_rows():
        row_cells = []
        for cell in rows:
            row_cells.append(cell.value)
        l_cells.append(tuple(row_cells))
    
    i_cells = [] # movable y/n
    for rows in crossover_inclusion.iter_rows():
        row_cells = []
        for cell in rows:
            row_cells.append(cell.value)
        i_cells.append(tuple(row_cells))
    
    land_code = {r:list() for r in range(11)}
    chromosome = {r:list() for r in range(11)}
    
    for m,n in product(range(idx),range(cols)):
        
        code = l_cells[m][n]
        land_code[code].append((m,n))
        
        if i_cells[m][n] == 1:
            chromosome[code].append((m,n))

    code_len = [len(c) for c in chromosome.values()]
    
    vars_dct = {'transport':transportation, 'flooding':flood_ratio, 
                'center':center_pt, 'code':land_code, 'chromosome':chromosome, 'code_len':code_len}
    cells_dct = {'n': n_cells, 'l':l_cells, 'i':i_cells}
    
    return vars_dct, cells_dct

def totalFitness(vars_dct, cells_dct):
    social, social_score = socialFitness(vars_dct['code'], vars_dct['chromosome'], vars_dct['transport'], cells_dct['n'])
    comm, comm_score = commFitness(vars_dct['chromosome'], vars_dct['transport'])
    env, env_score = envFitness(vars_dct['chromosome'], vars_dct['flooding'], vars_dct['center'], cells_dct['n'])
    
    return social, comm, env, [social_score, comm_score, env_score]
    
    

if __name__ == "__main__":
    # (n,l,i) = (flood risk, land code, inclusion)
    # n = 1: high flood risk     # n = 0: low flood risk   # n = -1: n/a
    # i = 1: movable             # i = 0: permanent 
    # l = 0: Non-Hoboken/ignore  # l = 1: water            # l = 2: undecided    
    # l = 3: university          # l = 4: industrial       # l = 5: residential
    # l = 6: recreation          # l = 7: hospital         # l = 8: fire station       
    # l = 9: commercial          # l = 10: waterfront

    # read map file
    fpath = "map_layers.xlsx"
    
    vars_dct, cells_dct = initVars(fpath)
    
    social, comm, env, total_score = totalFitness(vars_dct, cells_dct) # overall fitness
    
    # plotData(cells_dct['l']) # show map of current cells